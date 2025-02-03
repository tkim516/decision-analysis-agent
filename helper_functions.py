from langchain_openai import OpenAIEmbeddings
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_pinecone import PineconeVectorStore
from pinecone import Pinecone
import os
from langchain.schema import Document  # Import the Document class
from langchain_openai import OpenAIEmbeddings
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_pinecone import PineconeVectorStore
from pinecone import Pinecone
from langchain.schema import Document  # Import the Document class
import os
from openpyxl import load_workbook
import pandas as pd
from langchain.prompts import PromptTemplate
from langchain_openai import ChatOpenAI


def initialize_vector_store():
    embeddings = OpenAIEmbeddings(
        model="text-embedding-3-large",
        openai_api_key=os.environ.get("OPENAI_API_KEY")
    )

    pc = Pinecone(api_key=os.environ.get('PINECONE_API_KEY'))
    index_name = "decision-analysis-embeddings"
    index = pc.Index(index_name)
    vector_store = PineconeVectorStore(embedding=embeddings, index=index)
   
    return vector_store


def add_problem_embeddings(question_text: str, question_id: str, vector_store):

    text_splitter = RecursiveCharacterTextSplitter(
        chunk_size=10000, 
        chunk_overlap=200,
        add_start_index=True
    )
    
    # 1) First, wrap your single question string into a Document
    initial_docs = [Document(page_content=question_text)]
    
    # 2) Now, split those Documents
    all_splits = text_splitter.split_documents(initial_docs)
    
    # 3) Update metadata for each chunk
    documents = [
        Document(
            page_content=doc.page_content, 
            metadata={"question_id": question_id, "chunk_index": i}
        )
        for i, doc in enumerate(all_splits)
    ]

    _ = vector_store.add_documents(documents=documents)


def retrieve_similar_questions(target_question: str, vector_store, k=1):
    retrieved_docs = vector_store.similarity_search(target_question, k=k)
    
    # Build a list of results, each with question_id and content
    results = []
    for doc in retrieved_docs:
        # Pull the question_id from the metadata (default "N/A" if missing)
        similar_question_id = doc.metadata.get("question_id", "N/A")
        similar_question_text = doc.page_content
    
    return similar_question_id, similar_question_text


def get_answer_text(question_id: str, question_set):
    answer_text = question_set[question_set['questionId'] == question_id]['answerText'].values[0]
    return answer_text


def parse_excel_data(question_id: str, question_set):
    excel_file_name = question_set[question_set['questionId'] == question_id]['excelFileName'].values[0]
    
    # Load the workbook
    file_path = f'/Users/tyler/Downloads/ML/decision-analysis-agent/ExcelAnswers/{excel_file_name}'
    wb = load_workbook(file_path, data_only=False)  # Set data_only=False to preserve formulas

    # Select the worksheet
    sheet = wb.active

    # Extract data and formulas
    data = []
    formulas = []

    for row in sheet.iter_rows():
        row_data = []
        #row_formulas = []
        for cell in row:
            row_data.append(cell.value)  # This retrieves the value (result of formula)
            #row_formulas.append(cell.formula if hasattr(cell, 'formula') else None)  # This retrieves the formula
        data.append(row_data)
        #formulas.append(row_formulas)

    #print("Formulas:", formulas)
        
    return data


def extract_excel_logic(similar_question_text: str, answer_text: str, excel_data: list, llm):

    prompt = PromptTemplate(
    input_variables=[
        'similar_question_text',
        'answer_text',
        'excel_data'],

    template="""
    You are an assistant that extracts the logic in Excel that is used to solve decision analysis problems.

    Here is the question: {similar_question_text}

    Here is a list of lists which represents the Excel model used to solve the above question: {excel_data}

    Here is an explanation of the answer: {answer_text}

    Please explain the logic used to solve this problem. Also, give instructions for how another LLM can use this logic to solve similar problems.
    """
  )

    message = prompt.invoke({
        'similar_question_text': similar_question_text,
        'answer_text': answer_text,
        'excel_data': excel_data})
    
    response = llm.invoke(message)

    return response


def solve_problem(target_question: str, logic_instructions, llm):
    
    prompt = PromptTemplate(
    input_variables=[
        'target_question'
        'logic_instructions'],

    template="""
    You are an assistant that answers decision analysis problems.

    Answer this question: {target_question}

    Use this logical framework derived from a similar problem as a starting point: {logic_instructions}
    """
    )

    message = prompt.invoke({
        'target_question': target_question,
        'logic_instructions': logic_instructions})
    
    response = llm.invoke(message)

    return response